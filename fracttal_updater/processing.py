"""
Data processing module for reading Excel files and calculating meter values.
"""

import pandas as pd
from pathlib import Path


import openpyxl

def load_excel(filepath: str | Path) -> pd.DataFrame:
    """
    Load the RSV activity report Excel file.

    Args:
        filepath: Path to the Excel file

    Returns:
        DataFrame with the activity data
    """
    # The RSV report has 8 header rows to skip
    # Data starts at row 9 (header) and 10 (first data row) 
    df = pd.read_excel(filepath, header=8)
    
    # Standardize column names (strip whitespace)
    df.columns = df.columns.str.strip()
    
    return df


def mark_status(filepath: str | Path, row_index: int, status: str = "OK") -> None:
    """
    Mark the status of a specific row in the Excel file.

    Args:
        filepath: Path to the Excel file
        row_index: The DataFrame index of the row (0-based)
        status: The status text to write (default: "OK")
    """
    filepath = Path(filepath)
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        # Determine header row (1-based)
        # pd.read_excel(header=8) means the header is at row 9
        header_row = 9
        
        # Find "Estado" column index
        estado_col_idx = None
        for col in range(1, sheet.max_column + 2):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value and str(cell_value).strip() == "Estado":
                estado_col_idx = col
                break
        
        # If "Estado" column doesn't exist, create it
        if estado_col_idx is None:
            estado_col_idx = sheet.max_column + 1
            sheet.cell(row=header_row, column=estado_col_idx).value = "Estado"
            
        # Write status to the specific row
        # DataFrame index 0 corresponds to Excel row: header_row + 1 + index
        excel_row = header_row + 1 + row_index
        sheet.cell(row=excel_row, column=estado_col_idx).value = status
        
        wb.save(filepath)
        
    except Exception as e:
        print(f"Error updating status in Excel: {e}")


def calculate_value_to_add(row: pd.Series) -> tuple[float, str]:
    """
    Calculate the value to add based on the asset category.

    Args:
        row: A row from the DataFrame

    Returns:
        Tuple of (value_to_add, unit_type)
        unit_type is 'Km' for vehicles or 'Horas' for machinery
    """
    categoria = str(row.get("Categoría", "")).strip()

    if categoria in ["Flota Liviana", "Camiones"]:
        # Use Km column
        try:
            valor = float(str(row.get("Km", 0)).replace(",", "."))
            return valor, "Km"
        except ValueError:
            return 0.0, "Km"

    elif categoria == "Maquinarias":
        # Use "Tiempo de marcha" (HH:MM -> decimal hours)
        tiempo = str(row.get("Tiempo de marcha", "0:00")).strip()
        try:
            parts = tiempo.split(":")
            horas = int(parts[0])
            minutos = int(parts[1]) if len(parts) > 1 else 0
            return horas + minutos / 60, "Horas"
        except (ValueError, IndexError):
            return 0.0, "Horas"

    else:
        # Unknown category
        return 0.0, "Desconocido"


def get_interno_and_categoria(row: pd.Series) -> tuple[str, str]:
    """
    Extract the internal ID and category from a row.

    Args:
        row: A row from the DataFrame

    Returns:
        Tuple of (interno, categoria)
    """
    interno = str(row.get("Interno", "")).strip()
    categoria = str(row.get("Categoría", "")).strip()
    return interno, categoria
